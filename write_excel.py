import openpyxl

# 打开 Excel 文件
workbook = openpyxl.load_workbook('./Server configuration.xlsx')

# 选择要操作的表格
sheet = workbook.active

# 选择要读取的行数
row_num = 2

# 读取指定行的单元格内容
a_value = sheet.cell(row=row_num, column=1).value
b_value = sheet.cell(row=row_num, column=2).value
c_value = sheet.cell(row=row_num, column=3).value
d_value = sheet.cell(row=row_num, column=4).value
e_value = sheet.cell(row=row_num, column=5).value
f_value = sheet.cell(row=row_num, column=6).value

# 将单元格内容组合成一句话
sentence = f'docker run -dit --gpus all --name {a_value} -p {b_value}:6006 -p {c_value}:22 -p {d_value}:8888 -v ~/ws4/{e_value}:/workspace {f_value}'

# 将句子写入 Excel 文件的指定位置
sheet.cell(row=row_num, column=8).value = sentence

# 保存 Excel 文件
workbook.save('./Server configuration.xlsx')